"""
ONE-TIME setup script. Reads the two assignment xlsx files and creates
matching boards + columns on Monday.com, then imports every row as an item.

This is the ONLY place raw file data is touched - the running agent
always queries Monday.com live, never these files.

Usage:
    1. Put Deal funnel Data.xlsx and Work_Order_Tracker Data.xlsx in this
       backend/ folder (or edit the paths below).
    2. export MONDAY_API_TOKEN=xxxx   (or put it in a .env file)
    3. python seed_monday.py
    4. Copy the printed board IDs into your .env file.
"""
import os
from datetime import date, datetime
from pathlib import Path
import time
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

from monday_client import MondayClient  # noqa: E402

BASE_DIR = Path(__file__).resolve().parent
DEALS_FILE = BASE_DIR / "Deal funnel Data.xlsx"
WO_FILE = BASE_DIR / "Work_Order_Tracker Data.xlsx"

def col_type_for(colname):
    lc = colname.lower()
    if "date" in lc:
        return "date"
    if any(k in lc for k in ["value", "amount", "quantity", "qty"]):
        return "numbers"
    return "text"


def build_board(monday, board_name, columns):
    board_id = monday.create_board(board_name)
    print(f"Created board '{board_name}' -> id {board_id}")
    col_ids = {}
    name_col = columns[0]  # first column becomes the item's built-in "name"
    for c in columns:
        if c == name_col:
            continue
        ctype = col_type_for(c)
        try:
            cid = monday.create_column(board_id, c, ctype)
        except Exception as e:
            print(f"  column '{c}' ({ctype}) failed ({e}), retrying as text")
            cid = monday.create_column(board_id, c, "text")
            ctype = "text"
        col_ids[c] = (cid, ctype)
        time.sleep(0.25)
    return board_id, name_col, col_ids


def is_blank(value):
    return value is None or value == ""


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    return datetime.fromisoformat(str(value)).strftime("%Y-%m-%d")


def row_to_column_values(row, col_ids):
    values = {}
    for c, (cid, ctype) in col_ids.items():
        v = row.get(c)
        if is_blank(v):
            continue
        if ctype == "date":
            try:
                values[cid] = {"date": format_date(v)}
            except Exception:
                pass
        elif ctype == "numbers":
            try:
                values[cid] = str(float(v))
            except Exception:
                pass
        else:
            values[cid] = str(v)[:2000]
    return values


def read_excel_rows(file_path, header_row=0):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    headers = list(rows[header_row])
    data = []
    for values in rows[header_row + 1:]:
        if all(is_blank(value) for value in values):
            continue
        row = dict(zip(headers, values))
        if row.get(headers[0]) == headers[0]:
            continue
        data.append(row)
    return headers, data


def seed(file_path, board_name, header_row=0):
    columns, rows = read_excel_rows(file_path, header_row)

    monday = MondayClient()
    board_id, name_col, col_ids = build_board(monday, board_name, columns)

    total = len(rows)
    for i, row in enumerate(rows):
        item_name = str(row[name_col]) if not is_blank(row.get(name_col)) else f"Row {i}"
        cvs = row_to_column_values(row, col_ids)
        try:
            monday.create_item(board_id, item_name, cvs)
        except Exception as e:
            print(f"  item '{item_name}' failed: {e}")
        if i % 20 == 0:
            print(f"  ...{i}/{total} rows")
        time.sleep(0.2)
    print(f"Done: {board_name} -> board_id = {board_id} ({total} rows)\n")
    return board_id


if __name__ == "__main__":
    assert os.environ.get("MONDAY_API_TOKEN"), "Set MONDAY_API_TOKEN first"

    deals_id = seed(DEALS_FILE, "Deals - Sales Pipeline", header_row=0)
    # The Work Order sheet has a blank first row before the real header row
    wo_id = seed(WO_FILE, "Work Orders - Execution", header_row=1)

    print("Add these to your .env file:")
    print(f"MONDAY_DEALS_BOARD_ID={deals_id}")
    print(f"MONDAY_WORKORDERS_BOARD_ID={wo_id}")
